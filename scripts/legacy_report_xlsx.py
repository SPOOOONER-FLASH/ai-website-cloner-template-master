"""
Builds the client-facing Excel report on the legacy-site migration.

Input is docs/research/legacy/report-data.json, derived from the scrape. Regenerate the
whole chain with:

    node scripts/scrape-legacy-products.mjs 1058 387 1597 1601 391 506
    node scripts/merge-legacy-specs.mjs --write
    node scripts/import-legacy-images.mjs --write
    py scripts/legacy_report_xlsx.py

The workbook is written for a non-technical Chinese-speaking reader, so every sheet
leads with what the client has to DO, not with what the pipeline found.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = json.loads((ROOT / "docs/research/legacy/report-data.json").read_text("utf-8"))
OUT = ROOT / "docs/research/legacy/旧站素材迁移报告.xlsx"

# A Latin face renders Chinese via fallback and the columns end up ragged, so the
# report uses a CJK face throughout.
FONT = "Microsoft YaHei"

TITLE = Font(name=FONT, size=14, bold=True)
HEAD = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
BODY_RED = Font(name=FONT, size=10, color="C00000")
NOTE = Font(name=FONT, size=9, color="595959", italic=True)
LINK = Font(name=FONT, size=9, color="0563C1", underline="single")

FILL_HEAD = PatternFill("solid", fgColor="1F3864")
FILL_BAD = PatternFill("solid", fgColor="FCE4E4")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def sheet(wb, name, title, subtitle, headers, widths):
    """Every sheet opens with a title, a one-line brief, then a frozen header row."""
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = TITLE
    ws["A2"] = subtitle
    ws["A2"].font = NOTE

    for i, (head, width) in enumerate(zip(headers, widths), start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=4, column=i, value=head)
        cell.font = HEAD
        cell.fill = FILL_HEAD
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"
    return ws


def write_row(ws, row, values, fill=None, red_cols=()):
    for i, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = BODY_RED if i in red_cols else BODY
        cell.border = BOX
        cell.alignment = WRAP
        if fill:
            cell.fill = fill
    return row + 1


wb = Workbook()
wb.remove(wb.active)

# ── 1. Images the client has to replace ──────────────────────────────────────────
imgs = DATA["needReshoot"]
ws = sheet(
    wb,
    "①需重新提供的图片",
    "需要甲方重新提供的产品图片",
    "旧站 2022 年上传的这批图，产品主体上压着对角线域名水印，无法去除，因此未采用。请甲方提供无水印原图。",
    ["型号", "产品名称", "图片序号", "问题", "水印内容", "旧站图片地址", "旧站产品页"],
    [12, 34, 10, 26, 20, 52, 46],
)

row = 5
for item in imgs:
    problem = "对角线域名水印 + 左上角徽标"
    # 600's watermark advertises a different brand entirely — call that out separately.
    if item["wm"] == "www.hydeland.cn":
        problem = "水印是【其他品牌域名】，最严重"
    row = write_row(
        ws,
        row,
        [
            item["model"],
            item["name"],
            f"第 {item['seq']} 张",
            problem,
            item["wm"],
            item["url"],
            item["legacy"],
        ],
        fill=FILL_BAD,
        red_cols=(4, 5),
    )

for r in range(5, row):
    for c in (6, 7):
        cell = ws.cell(row=r, column=c)
        cell.font = LINK
        cell.hyperlink = cell.value

total = row
ws.cell(row=total + 1, column=1, value="合计").font = Font(name=FONT, size=10, bold=True)
count = ws.cell(row=total + 1, column=3, value=f"=COUNTA(C5:C{row - 1})")
count.font = Font(name=FONT, size=10, bold=True)
ws.cell(
    row=total + 3,
    column=1,
    value="说明：这三个型号的规格文字已经采用，只有图片需要重新提供。若甲方能给出无水印原图，"
    "直接放进 CMS 后台的「图库」字段即可，不需要改代码。",
).font = NOTE

# ── 2. Field conflicts needing a client decision ─────────────────────────────────
ws = sheet(
    wb,
    "②材质字段冲突",
    "材质 / 表面处理 / 适用门型 —— 两个来源对不上",
    "左边是我们网站现在写的（来自微信素材包），右边是甲方旧站自己写的。两边都是甲方给的，需要甲方确认以哪个为准。",
    ["型号", "字段", "网站现在写的", "旧站写的", "严重程度", "需要甲方确认"],
    [12, 14, 40, 46, 14, 34],
)

FIELD_CN = {"material": "材质", "finishes": "表面处理", "doorTypes": "适用门型"}


def severity(field, ours, legacy):
    """Iron vs steel is a different metal, not a different wording — flag those apart."""
    if field != "material":
        return "措辞差异", FILL_WARN
    a, b = ours.lower(), legacy.lower()
    iron_a, iron_b = "iron" in a, "iron" in b
    if iron_a != iron_b:
        return "⚠ 材质不同", FILL_BAD
    return "措辞差异", FILL_WARN


row = 5
for c in DATA["conflicts"]:
    label, fill = severity(c["field"], c["ours"], c["legacy"])
    ask = "铁 / 钢 是两种金属，必须确认" if fill is FILL_BAD else "确认用哪个说法"
    row = write_row(
        ws,
        row,
        [c["model"], FIELD_CN.get(c["field"], c["field"]), c["ours"], c["legacy"], label, ask],
        fill=fill,
        red_cols=(5,),
    )

ws.cell(
    row=row + 2,
    column=1,
    value="这些字段本次一律未改动，网站上仍是原来的值。采购商会照着材质下单，所以由甲方拍板，不由我们推断。",
).font = NOTE

# ── 3. Certification claims that were blocked ────────────────────────────────────
ws = sheet(
    wb,
    "③认证待核实",
    "旧站上的认证声明 —— 已拦截，未采用",
    "认证是对外承诺，必须有一份点名该型号的检测报告才能publish。以下两条都缺报告支撑。",
    ["型号", "旧站写的", "问题", "处理", "需要甲方提供"],
    [12, 24, 46, 20, 40],
)

CLAIM_NOTE = {
    "EN1205": "该标准号不存在。逃生推杠的标准是 EN 1125，旧站同一页的 meta 里又写成 EN12205，明显是打错的。",
    "EN1125": "标准号本身有效，但我们手上的四份报告点名的是 KD070/30-290、KD070/20-101、607 SS ET，没有一份覆盖此型号。",
}

row = 5
for c in DATA["claims"]:
    row = write_row(
        ws,
        row,
        [
            c["model"],
            f"{c['label']}：{c['value']}",
            CLAIM_NOTE.get(c["value"], "缺报告支撑"),
            "未采用",
            "点名该型号的检测报告扫描件",
        ],
        fill=FILL_BAD,
        red_cols=(4,),
    )

ws.cell(
    row=row + 2,
    column=1,
    value="拿到报告后，在 CMS 后台该产品的「认证」字段里填写，并把报告文件传到「下载中心」关联上即可。",
).font = NOTE

# ── 4. What this round actually delivered ────────────────────────────────────────
ws = sheet(
    wb,
    "④本次已完成",
    "本次从旧站补回的内容",
    "以下内容已经上线到 spoonercantonlock.stahlock.com，构建通过。",
    ["型号", "产品名称", "新增图片", "规格行数", "图片来源批次", "状态"],
    [12, 40, 12, 12, 18, 16],
)

BATCH = {"023 ETAN": "2024 年", "314": "2026 年", "317": "2026 年"}
row = 5
for d in DATA["done"]:
    row = write_row(
        ws,
        row,
        [
            d["model"],
            d["name"],
            d["gallery"],
            d["specs"],
            BATCH.get(d["model"], "—"),
            "✅ 已上线",
        ],
        fill=FILL_OK,
    )

for item in ("305", "320", "600"):
    row = write_row(
        ws,
        row,
        [item, "（见工作表①）", 0, "已补规格", "2022 年 — 未采用", "⚠ 待甲方给图"],
        fill=FILL_WARN,
        red_cols=(6,),
    )

ws.cell(row=row + 1, column=1, value="合计新增图片").font = Font(name=FONT, size=10, bold=True)
tot = ws.cell(row=row + 1, column=3, value=f"=SUM(C5:C{row - 1})")
tot.font = Font(name=FONT, size=10, bold=True)

ws.cell(
    row=row + 3,
    column=1,
    value="另外：旧站共有 424 个产品，我们网站目前 20 个。是否继续批量导入，等甲方定。",
).font = NOTE

# ── 5. Products carrying no photography at all ──────────────────────────────────
no_image = DATA.get("noImage", [])
if no_image:
    ws = sheet(
        wb,
        "⑤缺图产品清单",
        f"没有任何图片的产品 —— 共 {len(no_image)} 个",
        "这些产品在旧站上只有 2022 那批水印图，按约定未采用，所以详情页现在显示占位方块。"
        "这是拍照或重新导出的工作清单，可直接按分类分批处理。",
        ["型号", "产品名称", "分类", "旧站有几张水印图", "旧站产品页"],
        [22, 30, 26, 18, 62],
    )

    row = 5
    # Group by category so the client can brief a shoot one category at a time.
    for item in sorted(no_image, key=lambda x: (x["category"], x["model"])):
        row = write_row(
            ws,
            row,
            [
                item["model"],
                item["name"],
                item["category"],
                item["dropped"] or "—",
                item["legacyUrl"],
            ],
            fill=FILL_WARN,
        )

    for r in range(5, row):
        cell = ws.cell(row=r, column=5)
        if cell.value:
            cell.font = LINK
            cell.hyperlink = cell.value

    ws.cell(row=row + 1, column=1, value="合计").font = Font(name=FONT, size=10, bold=True)
    total_cell = ws.cell(row=row + 1, column=2, value=f"=COUNTA(A5:A{row - 1})")
    total_cell.font = Font(name=FONT, size=10, bold=True)

    ws.cell(
        row=row + 3,
        column=1,
        value="两条路：① 甲方提供无水印原图（最快，图本来就存在，只是被水印毁了）；"
        "② 重新拍摄。建议先做前 100 个高频型号，不必一次拍完 282 个。",
    ).font = NOTE

# ── 6. Products with no specification table ─────────────────────────────────────
no_spec = DATA.get("noSpec", [])
if no_spec:
    ws = sheet(
        wb,
        "⑥缺规格产品",
        f"规格表为空的产品 —— 共 {len(no_spec)} 个",
        "旧站这些页面本身就没有规格行。详情页显示「尺寸待确认」的空状态，"
        "没有从相似产品推断任何数值。",
        ["型号", "产品名称", "分类", "需要甲方提供"],
        [22, 34, 26, 40],
    )

    row = 5
    for item in sorted(no_spec, key=lambda x: (x["category"], x["model"])):
        row = write_row(
            ws,
            row,
            [item["model"], item["name"], item["category"], "材质、尺寸、表面处理"],
            fill=FILL_WARN,
        )

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"→ {OUT.relative_to(ROOT)}")
print(f"  ①需重新提供的图片 {len(imgs)} 张")
print(f"  ②材质字段冲突 {len(DATA['conflicts'])} 处")
print(f"  ③认证待核实 {len(DATA['claims'])} 条")
print(f"  ④本次已完成 {len(DATA['done'])} 个产品")
