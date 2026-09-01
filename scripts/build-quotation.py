#!/usr/bin/env python3
"""Build a HYDE quotation workbook from a JSON job file.

    py scripts/build-quotation.py docs/quotations/<job>.json [-o <out.xlsx>]

Why this exists
---------------
The 2026-05-26 quotation was a hand-edited workbook: the logo, the header block and
the terms were baked into cells, so quoting a different buyer meant retyping the
company block and hoping nobody broke a merge. Product specifications were retyped by
hand too, which is how a quote drifts away from what the catalogue actually says.

This script keeps the layout in one place and takes everything that changes -- buyer,
terms, item list -- from a small JSON job file. Specifications come from
`content/products/<slug>.json`, so the quote and the website cannot disagree, and each
model links back to its product page.

Prices are optional. A job with no `unitPrice` on its items renders the price columns
empty but present, which is the "specifications now, prices to follow" quote sales
sends first.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
CACHE = ROOT / "tmp" / "claude-quotation" / "assets"

# --- design tokens -------------------------------------------------------------
INK = "111111"
MUTED = "6B6B6B"
RULE = "D0D0D0"
HEAD_BG = "1A1A1A"
ZEBRA = "FAFAFA"
FONT = "Arial"

# Column widths are chosen so the sheet prints on ONE A4 portrait page without Excel
# having to shrink it into illegibility. Excel's pixel width for a column is
# round(chars * 7) + 5, so the ten columns below come to ~830px against the ~746px of
# printable width left by 0.25" side margins -- a ~90% print scale. Widen anything here
# and the print scale drops with it.
COLS = {  # column letter -> width in characters
    "A": 1.5,
    "B": 4.0,    # No.
    "C": 13.0,   # Model + link
    "D": 22.0,   # Picture
    "E": 36.0,   # Specification
    "F": 6.5,    # Qty
    "G": 6.0,    # Unit
    "H": 11.0,   # Unit price
    "I": 12.0,   # Amount
    "J": 1.5,
}
FIRST, LAST = "B", "I"
PICTURE_COL_PX = 159   # D width in pixels, used to centre the product image
ROW_PX = 165           # minimum product row height in pixels
SPEC_CHARS_PER_LINE = 54  # characters that fit one line of column E at 8pt


def px_to_pt(px: float) -> float:
    return px * 0.75


def col_index(letter: str) -> int:
    return column_index_from_string(letter)


def content_cols() -> list[str]:
    return [get_column_letter(i) for i in range(col_index(FIRST), col_index(LAST) + 1)]


def thin(color: str = RULE) -> Side:
    return Side(style="thin", color=color)


def cell(ws, ref, value=None, *, size=10, bold=False, color=INK, italic=False,
         align="left", valign="center", wrap=False, fill=None, border=None,
         indent=0, link=None):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic,
                  underline="single" if link else None)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap,
                            indent=indent)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if border:
        c.border = border
    if link:
        c.hyperlink = link
    return c


def span(ws, row, col_from, col_to):
    ws.merge_cells(f"{col_from}{row}:{col_to}{row}")
    return f"{col_from}{row}"


def rule(ws, row, *, color=INK, weight="medium"):
    """A horizontal rule drawn as a bottom border across the content columns."""
    side = Side(style=weight, color=color)
    for col in content_cols():
        ws[f"{col}{row}"].border = Border(bottom=side)


# --- assets --------------------------------------------------------------------

def render_png(source: Path, out: Path, *, width: int) -> Path:
    """Rasterise SVG/WebP/PNG to a PNG that Excel can embed.

    Excel reads neither SVG nor WebP reliably across versions, and the catalogue
    stores both. SVG goes through sharp (already a dependency of the site build);
    everything else goes through Pillow. Output is cached under tmp/.
    """
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.exists() and out.stat().st_mtime >= source.stat().st_mtime:
        return out

    if source.suffix.lower() == ".svg":
        node = shutil.which("node")
        if not node:
            sys.exit("node is required to rasterise the SVG logo")
        script = (
            "const sharp=require('sharp');"
            f"sharp({json.dumps(str(source))},{{density:600}})"
            f".resize({{width:{width}}}).png().toFile({json.dumps(str(out))})"
            ".then(()=>{},e=>{console.error(e);process.exit(1)});"
        )
        subprocess.run([node, "-e", script], cwd=ROOT, check=True)
        return out

    im = Image.open(source).convert("RGBA")
    im.thumbnail((width, width), Image.LANCZOS)
    canvas = Image.new("RGBA", im.size, (255, 255, 255, 255))
    canvas.paste(im, (0, 0), im)
    canvas.convert("RGB").save(out, "PNG")
    return out


def place_image(ws, path: Path, col_letter: str, row: int, *, box_px: int,
                cell_w_px: int, cell_h_px: int):
    """Anchor an image centred inside one cell, scaled to fit a box_px square."""
    img = XLImage(str(path))
    scale = min(box_px / img.width, box_px / img.height, 1.0)
    w, h = int(img.width * scale), int(img.height * scale)
    off_x = max((cell_w_px - w) // 2, 0)
    off_y = max((cell_h_px - h) // 2, 0)
    marker = AnchorMarker(col=col_index(col_letter) - 1, colOff=pixels_to_EMU(off_x),
                          row=row - 1, rowOff=pixels_to_EMU(off_y))
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)),
    )
    ws.add_image(img)


# --- content -------------------------------------------------------------------

def load_product(slug: str) -> dict:
    path = ROOT / "content" / "products" / f"{slug}.json"
    if not path.exists():
        sys.exit(f"no catalogue entry for {slug} ({path})")
    return json.loads(path.read_text(encoding="utf-8"))


def spec_lines(product: dict, item: dict) -> list[str]:
    """Specification bullets for one item.

    Default source is the catalogue entry, so a quote cannot contradict the website.
    A job may pass `specs` to replace that list outright -- the catalogue carries
    scraped artefacts (numbered labels, a "Color" and a "Color Options" row saying the
    same thing) that are fine on a product page and wrong on a document a buyer reads
    line by line. `extraSpecs` appends to whichever list is used.

    Every bullet has to fit the E column on one line at 8pt, so keep them short.
    """
    lines: list[str] = []
    if item.get("specs"):
        lines = [f"• {line}" for line in item["specs"]]
    else:
        seen = set()
        for spec in product.get("specs", []):
            label = re.sub(r"^\s*\d+\.\s*", "", str(spec.get("label", ""))).strip()
            value = str(spec.get("value", "")).strip()
            if not value or value.lower().replace(" ", "") in seen:
                continue
            seen.add(value.lower().replace(" ", ""))
            lines.append(f"• {label}: {value}" if label else f"• {value}")
    return lines + [f"• {line}" for line in item.get("extraSpecs", [])]


def spec_row_height(specs: list[str]) -> int:
    """Row height that actually clears the wrapped specification text.

    A fixed height silently clips the longest item -- 307 and 311 carry seven bullets
    where 310 carries five. The E column holds roughly 60 characters per line at 8pt,
    and a wrapped 8pt line occupies about 11px.
    """
    lines = sum(max(1, -(-len(spec) // SPEC_CHARS_PER_LINE)) for spec in specs)
    return max(ROW_PX, lines * 11 + 18)


def resolve_image(product: dict, item: dict) -> Path | None:
    src = item.get("image") or (product.get("heroImage") or {}).get("src")
    if not src:
        return None
    path = ROOT / "public" / src.lstrip("/")
    return path if path.exists() else None


def product_url(job: dict, product: dict) -> str:
    base = job["company"].get("siteUrl", "https://cantonlock.com").rstrip("/")
    category = "/".join(product.get("categoryPath", []))
    return f"{base}/products/{category}/{product['slug']}/"


# --- build ---------------------------------------------------------------------

def build(job: dict, out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = job.get("sheetName", "Quotation")
    ws.sheet_view.showGridLines = False

    for col, width in COLS.items():
        ws.column_dimensions[col].width = width

    company, buyer = job["company"], job["buyer"]
    meta, terms = job.get("meta", {}), job.get("terms", [])

    # -- letterhead ------------------------------------------------------------
    for r, px in ((1, 10), (2, 30), (3, 22), (4, 22), (5, 6)):
        ws.row_dimensions[r].height = px_to_pt(px)

    logo_src = ROOT / company["logo"].lstrip("/")
    place_image(ws, render_png(logo_src, CACHE / "logo.png", width=900), "B", 2,
                box_px=190, cell_w_px=210, cell_h_px=58)

    cell(ws, span(ws, 2, "F", LAST), "QUOTATION", size=22, bold=True, align="right")
    cell(ws, span(ws, 3, "F", LAST), f"Quotation No.   {meta.get('number', '')}",
         size=9, color=MUTED, align="right")
    cell(ws, span(ws, 4, "F", LAST), f"Date   {meta.get('date', '')}",
         size=9, color=MUTED, align="right")

    ws.row_dimensions[6].height = px_to_pt(16)
    ws.row_dimensions[7].height = px_to_pt(16)
    cell(ws, span(ws, 6, FIRST, "E"), company["legalName"], size=10, bold=True)
    cell(ws, span(ws, 7, FIRST, "E"), company["addressLine"], size=9, color=MUTED)
    cell(ws, span(ws, 6, "F", LAST), company["email"], size=9, color=MUTED, align="right")
    cell(ws, span(ws, 7, "F", LAST), company["siteUrl"], size=9, color=MUTED,
         align="right", link=company["siteUrl"])
    rule(ws, 7)

    # -- buyer / commercial terms ---------------------------------------------
    ws.row_dimensions[8].height = px_to_pt(12)
    head_row = 9
    ws.row_dimensions[head_row].height = px_to_pt(18)
    cell(ws, span(ws, head_row, FIRST, "D"), "TO", size=8, bold=True, color=MUTED)
    cell(ws, span(ws, head_row, "E", LAST), "TERMS", size=8, bold=True, color=MUTED)

    left = [buyer["name"]] + buyer.get("lines", [])
    right = [f"{t['label']}:  {t['value']}" for t in terms]
    for i in range(max(len(left), len(right))):
        r = head_row + 1 + i
        ws.row_dimensions[r].height = px_to_pt(16)
        if i < len(left):
            cell(ws, span(ws, r, FIRST, "D"), left[i], size=10 if i == 0 else 9,
                 bold=(i == 0), color=INK if i == 0 else MUTED)
        if i < len(right):
            cell(ws, span(ws, r, "E", LAST), right[i], size=9, color=MUTED)

    # -- item table ------------------------------------------------------------
    table_head = head_row + 1 + max(len(left), len(right)) + 1
    ws.row_dimensions[table_head - 1].height = px_to_pt(14)
    ws.row_dimensions[table_head].height = px_to_pt(26)
    headers = [("B", "NO."), ("C", "MODEL"), ("D", "PICTURE"), ("E", "SPECIFICATION"),
               ("F", "QTY"), ("G", "UNIT"), ("H", "UNIT PRICE\n(USD)"),
               ("I", "AMOUNT\n(USD)")]
    for col, label in headers:
        cell(ws, f"{col}{table_head}", label, size=8, bold=True, color="FFFFFF",
             fill=HEAD_BG, align="center", wrap=True)

    box = Border(left=thin(), right=thin(), top=thin(), bottom=thin())
    row = table_head
    for idx, item in enumerate(job["items"], start=1):
        row += 1
        product = load_product(item["slug"])
        specs = spec_lines(product, item)
        row_px = item.get("rowHeightPx") or spec_row_height(specs)
        ws.row_dimensions[row].height = px_to_pt(row_px)
        fill = ZEBRA if idx % 2 == 0 else None

        for col in content_cols():
            c = ws[f"{col}{row}"]
            c.border = box
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)

        cell(ws, f"B{row}", idx, size=9, color=MUTED, align="center", border=box,
             fill=fill)
        # The model cell doubles as the link. Spelling the URL out inside the
        # specification cell cost two or three wrapped lines on every row and read as
        # clutter; "View online" under the model number does the same job in one.
        url = product_url(job, product)
        title = item.get("title") or f"{product['model']} {product['name']}"
        c = cell(ws, f"C{row}", align="center", wrap=True, border=box, fill=fill)
        c.value = CellRichText(
            TextBlock(InlineFont(rFont=FONT, sz=13, b=True, color=INK), title),
            TextBlock(InlineFont(rFont=FONT, sz=7, color=MUTED, u="single"),
                      "\nView online ›"),
        )
        c.hyperlink = url

        image = resolve_image(product, item)
        if image:
            png = render_png(image, CACHE / f"{Path(image).stem}.png", width=560)
            place_image(ws, png, "D", row, box_px=min(item.get("imageBoxPx", 150), PICTURE_COL_PX - 8),
                        cell_w_px=PICTURE_COL_PX, cell_h_px=row_px)

        cell(ws, f"E{row}", "\n".join(spec_lines(product, item)), size=8, wrap=True,
             border=box, fill=fill, indent=1)

        cell(ws, f"F{row}", item.get("qty"), size=10, align="center", border=box,
             fill=fill)
        cell(ws, f"G{row}", item.get("unit", "PCS"), size=9, align="center", border=box,
             fill=fill)
        price = cell(ws, f"H{row}", item.get("unitPrice"), size=10, align="center",
                     border=box, fill=fill)
        price.number_format = '#,##0.00;;""'
        amount = cell(ws, f"I{row}", f'=IF(H{row}="","",F{row}*H{row})', size=10,
                      align="center", border=box, fill=fill)
        amount.number_format = '#,##0.00;;""'

    # -- total -----------------------------------------------------------------
    first_item, last_item = table_head + 1, row
    row += 1
    ws.row_dimensions[row].height = px_to_pt(26)
    top = Border(top=Side(style="medium", color=INK))
    cell(ws, span(ws, row, FIRST, "E"), "TOTAL", size=9, bold=True, align="right",
         border=top)
    for col in ("F", "G", "H"):
        ws[f"{col}{row}"].border = top
    cell(ws, f"H{row}", fill="F2F2F2")
    total = cell(ws, f"I{row}",
                 f'=IF(SUM(H{first_item}:H{last_item})=0,"",'
                 f'SUM(I{first_item}:I{last_item}))',
                 size=11, bold=True, align="center", border=top, fill="F2F2F2")
    total.number_format = '#,##0.00;;""'

    # -- notes -----------------------------------------------------------------
    row += 2
    cell(ws, span(ws, row, FIRST, LAST), "NOTES", size=8, bold=True, color=MUTED)
    for note in job.get("notes", []):
        row += 1
        ws.row_dimensions[row].height = px_to_pt(15)
        cell(ws, span(ws, row, FIRST, LAST), f"•  {note}", size=9)

    row += 2
    cell(ws, span(ws, row, FIRST, LAST), job.get("closing", ""), size=9, color=MUTED,
         italic=True)

    # -- print setup: A4 portrait, one page wide -------------------------------
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:J{row}"
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.25
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    out_path.parent.mkdir(parents=True, exist_ok=True)
    # openpyxl cannot overwrite a workbook Excel holds open; write elsewhere and move.
    tmp = Path(tempfile.mkdtemp()) / out_path.name
    wb.save(tmp)
    shutil.move(str(tmp), out_path)
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser(description="Build a HYDE quotation workbook.")
    ap.add_argument("job", type=Path, help="JSON job file")
    ap.add_argument("-o", "--out", type=Path,
                    help="output .xlsx (default: the job's outFile)")
    args = ap.parse_args()

    job = json.loads(args.job.read_text(encoding="utf-8"))
    out = args.out or Path(job["outFile"]).expanduser()
    print(f"wrote {build(job, out)}")


if __name__ == "__main__":
    main()
