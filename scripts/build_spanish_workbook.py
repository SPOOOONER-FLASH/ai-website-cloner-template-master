#!/usr/bin/env python
"""Turn the JSON payload from build-spanish-review-sheet.mjs into a translator's workbook.

Split in two on purpose. Node owns the catalogue: it knows how the Spanish glossary is
keyed and how many products carry each term, and it should not also be learning to write
a spreadsheet. Python owns the file format, because openpyxl is already installed here
and a hand-rolled xlsx writer is a liability nobody asked for.

The person who opens this file is a professional translator who has never seen the repo.
So the first sheet is instructions, every sheet has a single obvious column to type in,
and the sheets are ordered by consequence rather than alphabetically — twenty finish
terms that fix ninety-five strings come before seven hundred rows of spot-checking.

Usage:  py scripts/build_spanish_workbook.py <payload.json> [out.xlsx]
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# A face that carries Latin, CJK and the Spanish diacritics without silently substituting.
BODY = "Microsoft YaHei"

INK = "1A1917"
RULE = "D9D5CF"
HEAD_FILL = "1A1917"
ACCENT_FILL = "FBF3DF"   # the column the translator types in
GAP_FILL = "FDEDE9"      # rows that are visibly English on a Spanish page today

COLUMNS = [
    ("类型 / Tipo", 22),
    ("English", 46),
    ("Español (actual)", 40),
    ("中文（供 Spooner 核对）", 30),
    ("产品数 / N.º productos", 12),
    ("状态 / Estado", 26),
    ("✍ 译者修改 / Corrección", 40),
    ("备注 / Notas", 30),
]

SHEETS = [
    (
        "1 表面处理 Acabados",
        "finishes",
        "最高优先级。这二十条术语一次决定，替掉目录里 {finishSpellings} 种写法。"
        "PRIORIDAD MÁXIMA: estos 20 términos sustituyen {finishSpellings} variantes del catálogo.",
    ),
    (
        "2 缺译 Faltantes",
        "gaps",
        "这些词目前在西班牙语页面上仍显示英文。"
        "Estos términos aparecen HOY en inglés dentro de las páginas en español.",
    ),
    (
        "3 术语复核 Terminología",
        "terminology",
        "已有译法，按使用产品数从多到少排列 —— 先看上面的行，一行影响几百个页面。"
        "Ya traducido, ordenado por impacto: las primeras filas afectan cientos de páginas.",
    ),
    (
        "4 页面文案 Textos",
        "prose",
        "手写的整句文案：标题、导语、meta description。按句子审。"
        "Texto redactado a mano: títulos, entradillas y meta descriptions.",
    ),
    (
        "5 尺寸参考 Medidas",
        "dimensions",
        "纯尺寸与数字，无需翻译，仅供查证。"
        "Solo medidas y cifras. No requieren traducción; se incluyen como referencia.",
    ),
]

README = [
    ("这份表格要你做什么", "Qué se le pide"),
    (
        "本网站的西班牙语产品页不是翻译出来的，是用一份术语表拼出来的："
        "规格标签、规格值、产品类别各查一次，再组成西语句子。"
        "所以改一行术语，等于改掉所有用到它的页面。",
        "Las fichas de producto en español no se traducen frase a frase: se componen "
        "a partir de un glosario (etiquetas, valores y categorías). Corregir una fila "
        "corrige todas las páginas que usan ese término.",
    ),
    (
        "术语表里查不到的词，程序不会猜，会原样留英文。"
        "所以「缺译」那一页就是现在西语页面上真实可见的英文。",
        "Un término sin entrada en el glosario NO se adivina: se deja en inglés. "
        "Por eso la hoja «Faltantes» es lo que hoy se ve en inglés en el sitio.",
    ),
    (
        "只填「✍ 译者修改」一列。其他列请不要改动 —— 它们是程序读取的键。",
        "Escriba únicamente en la columna «✍ Corrección». No modifique las demás "
        "columnas: son las claves que lee el generador.",
    ),
    (
        "语域：拉丁美洲外贸西语（哥伦比亚、厄瓜多尔、秘鲁、阿根廷、墨西哥）。"
        "与西班牙半岛用法冲突时以拉美为准。",
        "Registro: español comercial latinoamericano (Colombia, Ecuador, Perú, "
        "Argentina, México). Ante divergencia con el uso peninsular, prevalece el "
        "latinoamericano.",
    ),
    (
        "重要：型号、尺寸、螺纹、中心距一律不翻译，也不要换算单位。"
        "买家按这些数字下单。",
        "IMPORTANTE: no traduzca ni convierta modelos, medidas, roscas ni distancias "
        "entre ejes. El comprador especifica a partir de esas cifras.",
    ),
]


def style_header(ws, row=1):
    for col, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(name=BODY, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = f"A{row + 1}"


def add_sheet(wb, title, note, rows):
    ws = wb.create_sheet(title[:31])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    head = ws.cell(row=1, column=1, value=note)
    head.font = Font(name=BODY, size=10, bold=True, color=INK)
    head.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    style_header(ws, row=2)

    thin = Side(style="thin", color=RULE)
    border = Border(bottom=thin)

    for i, r in enumerate(rows, start=3):
        values = [
            r.get("kind", ""),
            r.get("en", ""),
            r.get("es", ""),
            r.get("zh", ""),
            r.get("uses", ""),
            r.get("status", ""),
            "",
            "",
        ]
        is_gap = "缺译" in r.get("status", "")
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name=BODY, size=9, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=col in (2, 3, 4, 7, 8))
            cell.border = border
            if col == 7:
                cell.fill = PatternFill("solid", fgColor=ACCENT_FILL)
            elif is_gap and col <= 6:
                cell.fill = PatternFill("solid", fgColor=GAP_FILL)

    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{max(2, len(rows) + 2)}"
    return ws


def build(payload_path: Path, out_path: Path) -> Path:
    data = json.loads(payload_path.read_text(encoding="utf8"))
    meta = data.get("meta", {})

    wb = Workbook()
    wb.remove(wb.active)

    intro = wb.create_sheet("0 说明 Instrucciones")
    intro.column_dimensions["A"].width = 62
    intro.column_dimensions["B"].width = 72
    row = 1
    title = intro.cell(row=row, column=1, value="Canton Hyland · 西班牙语复核 / Revisión de español")
    title.font = Font(name=BODY, size=14, bold=True, color=INK)
    row += 2
    for zh, es in README:
        a = intro.cell(row=row, column=1, value=zh)
        b = intro.cell(row=row, column=2, value=es)
        bold = es == "Qué se le pide"
        for cell in (a, b):
            cell.font = Font(name=BODY, size=10, bold=bold, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        intro.row_dimensions[row].height = 24 if bold else 62
        row += 1

    row += 1
    counts = intro.cell(row=row, column=1, value="本次导出 / Este archivo")
    counts.font = Font(name=BODY, size=11, bold=True, color=INK)
    row += 1
    for sheet_title, key, _ in SHEETS:
        intro.cell(row=row, column=1, value=sheet_title).font = Font(name=BODY, size=10, color=INK)
        intro.cell(row=row, column=2, value=f"{len(data.get(key, []))} 行 / filas").font = Font(
            name=BODY, size=10, color=INK
        )
        row += 1
    intro.cell(row=row + 1, column=1, value=f"生成日期 / Generado: {meta.get('generated', '')}").font = Font(
        name=BODY, size=9, color="7A746C"
    )

    for sheet_title, key, note in SHEETS:
        add_sheet(wb, sheet_title, note.format(**meta), data.get(key, []))

    try:
        wb.save(out_path)
    except PermissionError:
        # Excel holds a lock on an open file; write beside it rather than losing the run.
        out_path = out_path.with_name(f"{out_path.stem}-new{out_path.suffix}")
        wb.save(out_path)
    return out_path


if __name__ == "__main__":
    payload = Path(sys.argv[1] if len(sys.argv) > 1 else "docs/collaboration/spanish-review.json")
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else payload.with_name("Canton-Hyland-西班牙语复核.xlsx")
    written = build(payload, out)
    print(f"wrote {written}")
